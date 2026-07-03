from datetime import date
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from flask_babel import gettext as _

HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(bold=True, size=14, color='1F4E79')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
ALT_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
TOTAL_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
TOTAL_FONT = Font(bold=True, color='FFFFFF', size=11)
CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(vertical='center')


def style_header(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_data_cell(cell, alt=False):
    cell.border = THIN_BORDER
    cell.alignment = LEFT
    if alt:
        cell.fill = ALT_FILL


def auto_width(ws, num_cols, max_rows=100):
    for col in range(1, num_cols + 1):
        max_len = 0
        for row in range(1, min(max_rows, ws.max_row or 1) + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(max_len + 2, 12), 30)


def export_vacation_plan(requests, filename_prefix='vacation_plan'):
    wb = Workbook()
    ws = wb.active
    ws.title = _('Vacation Plan')

    ws.merge_cells('A1:H1')
    title_cell = ws.cell(row=1, column=1, value=_('Vacation Plan Report'))
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:H2')
    date_cell = ws.cell(row=2, column=1, value=_('Generated on: %(date)s', date=date.today().strftime('%d/%m/%Y')))
    date_cell.font = Font(italic=True, color='666666', size=10)
    date_cell.alignment = Alignment(horizontal='center')

    headers = [
        _('Employee'), _('Department'), _('Start Date'), _('End Date'),
        _('Days'), _('Type'), _('Status'), _('Cause/Reason')
    ]
    header_row = 4
    for col, header in enumerate(headers, 1):
        ws.cell(row=header_row, column=col, value=header)
    style_header(ws, header_row, len(headers))

    for row_idx, req in enumerate(requests, header_row + 1):
        alt = (row_idx % 2 == 0)
        is_hr = req.request_type == 'hr_assigned'
        values = [
            req.user.display_name or req.user.username,
            req.user.department.name if req.user.department else '',
            req.start_date,
            req.end_date,
            req.days_count,
            _('HR Assigned') if is_hr else _('User Request'),
            _(req.status.capitalize()),
            req.cause.name if req.cause else (req.reason or ''),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col_idx in (3, 4):
                cell.number_format = 'DD/MM/YYYY'
            style_data_cell(cell, alt=alt)

    total_row = header_row + len(requests) + 1
    ws.cell(row=total_row, column=1, value=_('TOTAL')).font = TOTAL_FONT
    ws.cell(row=total_row, column=1).fill = TOTAL_FILL
    ws.cell(row=total_row, column=1).alignment = CENTER
    ws.cell(row=total_row, column=1).border = THIN_BORDER
    for col in range(2, len(headers) + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER
        if col == 5:
            cell.value = f'=SUM(E{header_row + 1}:E{total_row - 1})'
            cell.font = TOTAL_FONT
            cell.alignment = CENTER
            cell.number_format = '0'
        else:
            cell.font = Font(color='FFFFFF', size=11)

    avg_row = total_row + 1
    ws.cell(row=avg_row, column=1, value=_('AVERAGE')).font = Font(bold=True, size=10, color='1F4E79')
    ws.cell(row=avg_row, column=1).border = THIN_BORDER
    for col in range(2, len(headers) + 1):
        cell = ws.cell(row=avg_row, column=col)
        cell.border = THIN_BORDER
        if col == 5:
            cell.value = f'=AVERAGE(E{header_row + 1}:E{total_row - 1})'
            cell.number_format = '0.0'
            cell.font = Font(bold=True, size=10, color='1F4E79')
            cell.alignment = CENTER

    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f'A{header_row}:H{total_row - 1}'
    auto_width(ws, len(headers))

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
